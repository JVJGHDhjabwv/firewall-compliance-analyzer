import pandas as pd
import os
from datetime import date
from io import BytesIO
import argparse

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ============================================================
# Firewall Compliance Analyzer
# ============================================================

REQUIRED_COLUMNS = [
    "Rule_ID",
    "Rule_Name",
    "Source_Zone",
    "Source",
    "Source_IP",
    "Destination_Zone",
    "Destination",
    "Destination_IP",
    "Service",
    "Protocol",
    "Port",
    "Action",
    "Logging",
    "NAT",
    "Rule_Type",
    "Owner",
    "Department",
    "Business_Justification",
    "Ticket_Number",
    "Created_Date",
    "Last_Reviewed_Date",
    "Expiry_Date",
    "Status",
    "Environment",
    "Remarks"
]

SEVERITY_RANK = {
    "Critical": 4,
    "High": 3,
    "Medium": 2,
    "Low": 1
}

SEVERITY_ORDER = ["Critical", "High", "Medium", "Low"]

SEVERITY_COLORS = {
    "Critical": "C00000",
    "High":     "FF0000",
    "Medium":   "FF8C00",
    "Low":      "FFD700"
}

EXTERNAL_ZONES = {"WAN", "INTERNET", "EXTERNAL", "OUTSIDE"}
INTERNAL_ZONES = {"LAN", "INTERNAL", "INSIDE", "SERVER", "DB", "APP"}

TEMPORARY_RULE_TYPES = {"TEMP", "TEMPORARY", "VENDOR", "EMERGENCY", "TEST", "UAT"}

RISKY_SERVICES = {
    "RDP": {
        "ports": {"3389"},
        "severity": "High",
        "risk": (
            "RDP provides remote desktop access. If exposed broadly, attackers "
            "may attempt brute-force attacks or exploit remote access vulnerabilities."
        ),
        "recommendation": (
            "Restrict RDP to VPN or admin networks only and enable strong authentication."
        )
    },
    "SSH": {
        "ports": {"22"},
        "severity": "Medium",
        "risk": (
            "SSH provides remote administrative access. If exposed unnecessarily, "
            "it may be targeted for brute-force login attempts."
        ),
        "recommendation": "Restrict SSH access to trusted administrator IP ranges only."
    },
    "TELNET": {
        "ports": {"23"},
        "severity": "High",
        "risk": (
            "Telnet transmits data in clear text, including usernames and passwords. "
            "This can expose credentials to interception."
        ),
        "recommendation": "Replace Telnet with SSH and block Telnet access."
    },
    "FTP": {
        "ports": {"21"},
        "severity": "High",
        "risk": (
            "FTP can transmit credentials and data without encryption, "
            "making it unsuitable for sensitive environments."
        ),
        "recommendation": "Use SFTP or FTPS instead of FTP."
    },
    "SMB": {
        "ports": {"445"},
        "severity": "High",
        "risk": (
            "SMB is commonly abused for lateral movement and file-sharing attacks. "
            "Broad SMB access can increase ransomware impact."
        ),
        "recommendation": (
            "Restrict SMB to required internal systems only and avoid "
            "exposing it across network zones."
        )
    },
    "MSSQL": {
        "ports": {"1433"},
        "severity": "Medium",
        "risk": (
            "Database services should not be broadly reachable. "
            "Unrestricted database access can expose sensitive data."
        ),
        "recommendation": "Allow database access only from approved application servers."
    },
    "MYSQL": {
        "ports": {"3306"},
        "severity": "Medium",
        "risk": (
            "MySQL database access should be limited to authorised systems. "
            "Broad access may expose application data."
        ),
        "recommendation": "Restrict MySQL access to approved application servers only."
    },
    "POSTGRESQL": {
        "ports": {"5432"},
        "severity": "Medium",
        "risk": (
            "PostgreSQL database access should be limited to authorised systems. "
            "Broad access may expose application data."
        ),
        "recommendation": "Restrict PostgreSQL access to approved application servers only."
    },
    "MONGODB": {
        "ports": {"27017"},
        "severity": "Medium",
        "risk": (
            "MongoDB is frequently targeted due to misconfigured public exposure. "
            "Broad access may expose application or customer data."
        ),
        "recommendation": "Restrict MongoDB access to approved application servers only."
    },
    "REDIS": {
        "ports": {"6379"},
        "severity": "Medium",
        "risk": (
            "Redis has no authentication by default in older versions. "
            "Broad access may expose cached data or allow command injection."
        ),
        "recommendation": "Restrict Redis to internal application servers and enable authentication."
    },
    "LDAP": {
        "ports": {"389"},
        "severity": "High",
        "risk": (
            "LDAP without TLS exposes directory queries and credentials in clear text. "
            "Broad access increases the risk of directory enumeration."
        ),
        "recommendation": "Use LDAPS (636) and restrict access to authorised systems only."
    },
    "LDAPS": {
        "ports": {"636"},
        "severity": "Medium",
        "risk": (
            "LDAPS provides encrypted directory access but should still be restricted "
            "to authorised systems to prevent enumeration."
        ),
        "recommendation": "Restrict LDAPS access to approved identity-aware systems only."
    },
    "SNMP": {
        "ports": {"161"},
        "severity": "Medium",
        "risk": (
            "SNMP can expose device configuration and network topology information. "
            "SNMPv1 and v2c use community strings with no encryption."
        ),
        "recommendation": "Use SNMPv3 with authentication and restrict access to monitoring servers."
    },
    "SMTP": {
        "ports": {"25"},
        "severity": "Medium",
        "risk": (
            "Open SMTP access may be abused for spam relay or mail-based attacks "
            "if not restricted to authorised mail servers."
        ),
        "recommendation": "Restrict SMTP to authorised mail relay servers only."
    },
    "HTTP": {
        "ports": {"80"},
        "severity": "Low",
        "risk": (
            "HTTP traffic is not encrypted. Sensitive data transmitted over HTTP "
            "may be intercepted."
        ),
        "recommendation": "Use HTTPS instead of HTTP where possible."
    },
    "NETBIOS": {
        "ports": {"137", "138", "139"},
        "severity": "High",
        "risk": (
            "NetBIOS is an older protocol that can expose machine names, workgroup "
            "information, and file shares. It is frequently targeted in Windows environments."
        ),
        "recommendation": "Disable NetBIOS where not required and block it at zone boundaries."
    },
    "WINRM": {
        "ports": {"5985", "5986"},
        "severity": "High",
        "risk": (
            "WinRM provides remote management access to Windows systems. "
            "Broad access may allow remote command execution."
        ),
        "recommendation": "Restrict WinRM to authorised management stations and jump servers."
    },
    "VNC": {
        "ports": {"5900"},
        "severity": "High",
        "risk": (
            "VNC provides remote graphical access. Older versions may lack strong "
            "authentication or encryption."
        ),
        "recommendation": "Restrict VNC to trusted admin networks and use strong authentication."
    },
    "DOCKER": {
        "ports": {"2375", "2376"},
        "severity": "Critical",
        "risk": (
            "Exposed Docker API ports can allow full container and host compromise "
            "if accessible without authentication."
        ),
        "recommendation": "Block Docker API ports from all untrusted networks immediately."
    },
    "KUBERNETES": {
        "ports": {"6443", "10250"},
        "severity": "Critical",
        "risk": (
            "Exposed Kubernetes API or kubelet ports can allow cluster-level compromise "
            "and access to all workloads."
        ),
        "recommendation": "Restrict Kubernetes API and kubelet access to trusted management networks only."
    }
}

DUPLICATE_COLUMNS = [
    "Source_Zone",
    "Source",
    "Source_IP",
    "Destination_Zone",
    "Destination",
    "Destination_IP",
    "Service",
    "Protocol",
    "Port",
    "Action"
]


# ============================================================
# Helper Functions
# ============================================================

def is_empty(value):
    """
    Returns True if the value is blank, NaN, None, or a null-like string.
    """
    if pd.isna(value):
        return True

    return str(value).strip().lower() in ("", "nan", "none", "null", "-", "n/a")


def normalize(value):
    """
    Returns an uppercased, stripped string. Returns empty string for null values.
    """
    if pd.isna(value):
        return ""

    return str(value).strip().upper()


def get_value(rule, key):
    """
    Safely retrieves a rule field value with a fallback to empty string.
    """
    try:
        return rule[key]
    except (KeyError, TypeError):
        return ""


def load_firewall_rules(file_path):
    """
    Loads firewall rules from a CSV or Excel file.
    Strips whitespace from column names to handle formatting inconsistencies.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    extension = os.path.splitext(file_path)[1].lower()

    if extension == ".csv":
        df = pd.read_csv(file_path, dtype=str)
    elif extension in (".xlsx", ".xls"):
        df = pd.read_excel(file_path, dtype=str)
    else:
        raise ValueError(
            f"Unsupported file type '{extension}'. Use CSV, XLSX, or XLS."
        )

    df.columns = df.columns.astype(str).str.strip()
    df = df.reset_index(drop=True)

    return df


def validate_columns(df):
    """
    Checks that all required columns are present.
    Raises ValueError listing any missing columns.
    """
    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]

    if missing:
        raise ValueError(
            "Missing required columns: " + ", ".join(missing)
        )


def add_finding(rule, findings, issue, severity, risk_explanation, recommendation):
    """
    Appends a compliance finding to the findings list.
    """
    findings.append({
        "Rule_ID":               get_value(rule, "Rule_ID"),
        "Rule_Name":             get_value(rule, "Rule_Name"),
        "Source_Zone":           get_value(rule, "Source_Zone"),
        "Source":                get_value(rule, "Source"),
        "Source_IP":             get_value(rule, "Source_IP"),
        "Destination_Zone":      get_value(rule, "Destination_Zone"),
        "Destination":           get_value(rule, "Destination"),
        "Destination_IP":        get_value(rule, "Destination_IP"),
        "Service":               get_value(rule, "Service"),
        "Protocol":              get_value(rule, "Protocol"),
        "Port":                  get_value(rule, "Port"),
        "Action":                get_value(rule, "Action"),
        "Logging":               get_value(rule, "Logging"),
        "NAT":                   get_value(rule, "NAT"),
        "Rule_Type":             get_value(rule, "Rule_Type"),
        "Owner":                 get_value(rule, "Owner"),
        "Department":            get_value(rule, "Department"),
        "Business_Justification":get_value(rule, "Business_Justification"),
        "Ticket_Number":         get_value(rule, "Ticket_Number"),
        "Created_Date":          get_value(rule, "Created_Date"),
        "Last_Reviewed_Date":    get_value(rule, "Last_Reviewed_Date"),
        "Expiry_Date":           get_value(rule, "Expiry_Date"),
        "Status":                get_value(rule, "Status"),
        "Environment":           get_value(rule, "Environment"),
        "Remarks":               get_value(rule, "Remarks"),
        "Issue":                 issue,
        "Severity":              severity,
        "Risk_Explanation":      risk_explanation,
        "Recommendation":        recommendation
    })


# ============================================================
# Compliance Checks
# ============================================================

def check_any_source(rule, findings):
    action = normalize(get_value(rule, "Action"))
    source = normalize(get_value(rule, "Source"))
    source_ip = normalize(get_value(rule, "Source_IP"))

    if action == "ALLOW" and (source == "ANY" or source_ip == "ANY"):
        add_finding(
            rule, findings,
            "Source is set to Any",
            "High",
            "The rule allows traffic from any source. This increases the attack surface "
            "because unknown or unauthorised systems may reach the destination.",
            "Restrict the source to specific IP addresses, approved subnets, "
            "VPN ranges, or trusted network groups."
        )


def check_any_destination(rule, findings):
    action = normalize(get_value(rule, "Action"))
    destination = normalize(get_value(rule, "Destination"))
    destination_ip = normalize(get_value(rule, "Destination_IP"))

    if action == "ALLOW" and (destination == "ANY" or destination_ip == "ANY"):
        add_finding(
            rule, findings,
            "Destination is set to Any",
            "High",
            "The rule allows traffic to any destination. This may permit unnecessary "
            "access to systems that do not require connectivity.",
            "Restrict the destination to specific approved servers, applications, "
            "or network segments."
        )


def check_any_service(rule, findings):
    action = normalize(get_value(rule, "Action"))
    service = normalize(get_value(rule, "Service"))
    port = normalize(get_value(rule, "Port"))

    if action == "ALLOW" and (service == "ANY" or port == "ANY"):
        add_finding(
            rule, findings,
            "Service or port is set to Any",
            "Critical",
            "The rule allows all services or ports. This may expose unnecessary "
            "protocols and increase the chance of exploitation.",
            "Replace Any service or Any port with only the required ports and protocols."
        )


def check_any_any_any(rule, findings):
    action = normalize(get_value(rule, "Action"))
    source = normalize(get_value(rule, "Source"))
    source_ip = normalize(get_value(rule, "Source_IP"))
    destination = normalize(get_value(rule, "Destination"))
    destination_ip = normalize(get_value(rule, "Destination_IP"))
    service = normalize(get_value(rule, "Service"))
    port = normalize(get_value(rule, "Port"))

    source_any = source == "ANY" or source_ip == "ANY"
    destination_any = destination == "ANY" or destination_ip == "ANY"
    service_any = service == "ANY" or port == "ANY"

    if action == "ALLOW" and source_any and destination_any and service_any:
        add_finding(
            rule, findings,
            "Any-Any-Any Allow rule detected",
            "Critical",
            "This rule allows unrestricted traffic from anywhere to anywhere using any service. "
                        "It is one of the highest-risk firewall misconfigurations because it bypasses "
            "least-privilege access control.",
            "Remove the rule immediately or redesign it with specific source, "
            "destination, and service values."
        )


def check_missing_owner(rule, findings):
    if is_empty(get_value(rule, "Owner")):
        add_finding(
            rule, findings,
            "Missing rule owner",
            "Medium",
            "A rule without an owner has no clear accountability. This makes it difficult "
            "to confirm whether the rule is still required during firewall reviews.",
            "Assign a rule owner or responsible team."
        )


def check_missing_department(rule, findings):
    if is_empty(get_value(rule, "Department")):
        add_finding(
            rule, findings,
            "Missing department",
            "Low",
            "Without a department, it may be difficult to identify which business unit "
            "is responsible for the rule.",
            "Add the responsible department or business unit."
        )


def check_missing_business_justification(rule, findings):
    if is_empty(get_value(rule, "Business_Justification")):
        add_finding(
            rule, findings,
            "Missing business justification",
            "Medium",
            "Without a business justification, there is no documented reason explaining "
            "why the firewall rule is required.",
            "Add a clear business justification or remove the rule if it is not required."
        )


def check_missing_ticket_number(rule, findings):
    if is_empty(get_value(rule, "Ticket_Number")):
        add_finding(
            rule, findings,
            "Missing change ticket number",
            "Medium",
            "Without a change ticket number, the firewall rule may not be traceable "
            "to an approved change request.",
            "Add the related change request, service request, or approval ticket number."
        )


def check_missing_last_reviewed_date(rule, findings):
    if is_empty(get_value(rule, "Last_Reviewed_Date")):
        add_finding(
            rule, findings,
            "Missing last reviewed date",
            "Medium",
            "Rules without a review date may become outdated, unused, or overly "
            "permissive over time.",
            "Add the last reviewed date and include the rule in periodic firewall "
            "review activities."
        )


def check_logging_disabled(rule, findings):
    action = normalize(get_value(rule, "Action"))
    logging = normalize(get_value(rule, "Logging"))

    if action == "ALLOW" and logging in {"DISABLED", "NO", "FALSE", "OFF", ""}:
        add_finding(
            rule, findings,
            "Logging is disabled for allow rule",
            "Medium",
            "Without logging, allowed traffic may not be visible during monitoring, "
            "troubleshooting, or security investigations.",
            "Enable logging for allow rules, especially for high-risk or "
            "external-facing access."
        )


def check_rule_type_without_expiry(rule, findings):
    rule_type = normalize(get_value(rule, "Rule_Type"))
    expiry_date = get_value(rule, "Expiry_Date")

    if rule_type in TEMPORARY_RULE_TYPES and is_empty(expiry_date):
        add_finding(
            rule, findings,
            f"{rule_type.title()} rule has no expiry date",
            "High",
            "Temporary, vendor, emergency, or test rules may remain active longer "
            "than intended if no expiry date is defined.",
            "Add an expiry date so the rule can be reviewed, disabled, or removed "
            "after it is no longer required."
        )


def check_expired_rule(rule, findings):
    expiry_date = get_value(rule, "Expiry_Date")

    if is_empty(expiry_date):
        return

    parsed_date = pd.to_datetime(expiry_date, dayfirst=True, errors="coerce")

    if pd.isna(parsed_date):
        add_finding(
            rule, findings,
            "Invalid expiry date format",
            "Low",
            "An invalid expiry date prevents the system from determining whether "
            "the rule is still valid.",
            "Use a valid date format such as DD/MM/YYYY."
        )
        return

    today = pd.Timestamp(date.today())

    if parsed_date.normalize() < today:
        add_finding(
            rule, findings,
            "Rule has expired",
            "High",
            "The rule expiry date has passed, which means the access may no longer "
            "be approved or required.",
            "Review the rule and disable or remove it if the access is no longer needed."
        )


def check_expiry_within_30_days(rule, findings):
    """
    Warns when a rule is due to expire within 30 days.
    """
    expiry_date = get_value(rule, "Expiry_Date")

    if is_empty(expiry_date):
        return

    parsed_date = pd.to_datetime(expiry_date, dayfirst=True, errors="coerce")

    if pd.isna(parsed_date):
        return

    today = pd.Timestamp(date.today())
    days_remaining = (parsed_date.normalize() - today).days

    if 0 <= days_remaining <= 30:
        add_finding(
            rule, findings,
            f"Rule expiring soon ({days_remaining} days remaining)",
            "Medium",
            f"This rule is due to expire in {days_remaining} days. "
            "Rules that expire without review may be removed without notice.",
            "Review whether the access is still required and extend or remove "
            "the rule before the expiry date."
        )


def check_risky_services(rule, findings):
    action = normalize(get_value(rule, "Action"))

    if action != "ALLOW":
        return

    service = normalize(get_value(rule, "Service"))
    port = normalize(get_value(rule, "Port"))

    for risky_service, details in RISKY_SERVICES.items():
        if service == risky_service or port in details["ports"]:
            add_finding(
                rule, findings,
                f"Risky service detected: {risky_service}",
                details["severity"],
                details["risk"],
                details["recommendation"]
            )


def check_disabled_rule(rule, findings):
    status = normalize(get_value(rule, "Status"))

    if status == "DISABLED":
        add_finding(
            rule, findings,
            "Disabled rule detected",
            "Low",
            "Disabled rules may clutter the firewall policy and make rulebase "
            "reviews more difficult.",
            "Review disabled rules and remove them if they are no longer required."
        )


def check_external_to_internal(rule, findings):
    action = normalize(get_value(rule, "Action"))
    source_zone = normalize(get_value(rule, "Source_Zone"))
    destination_zone = normalize(get_value(rule, "Destination_Zone"))

    if (
        action == "ALLOW"
        and source_zone in EXTERNAL_ZONES
        and destination_zone in INTERNAL_ZONES
    ):
        add_finding(
            rule, findings,
            "External to internal allow rule",
            "High",
            "The rule allows traffic from an external zone into an internal zone. "
            "This may expose internal systems to external threats.",
            "Ensure the access is required, restrict the source, restrict the "
            "destination, limit the ports, and enable logging."
        )


def check_prod_any_rule(rule, findings):
    action = normalize(get_value(rule, "Action"))
    environment = normalize(get_value(rule, "Environment"))
    source = normalize(get_value(rule, "Source"))
    destination = normalize(get_value(rule, "Destination"))
    service = normalize(get_value(rule, "Service"))

    if environment == "PRODUCTION" and action == "ALLOW":
        if source == "ANY" or destination == "ANY" or service == "ANY":
            add_finding(
                rule, findings,
                "Production rule contains Any value",
                "High",
                "Production firewall rules with Any values are risky because they may "
                "expose critical systems or business services more broadly than required.",
                "Review the rule and apply least-privilege access by defining specific "
                "source, destination, and service values."
            )


def check_intrazone_allow(rule, findings):
    """
    Flags allow rules where source and destination zone are the same.
    These may indicate missing segmentation within a zone.
    """
    action = normalize(get_value(rule, "Action"))
    source_zone = normalize(get_value(rule, "Source_Zone"))
    destination_zone = normalize(get_value(rule, "Destination_Zone"))

    if (
        action == "ALLOW"
        and source_zone != ""
        and source_zone == destination_zone
    ):
        add_finding(
            rule, findings,
            "Intrazone allow rule detected",
            "Low",
            "The source and destination zone are the same. This may indicate a lack "
            "of micro-segmentation within the zone.",
            "Review whether intrazone traffic should be unrestricted or whether "
            "finer-grained segmentation is appropriate."
        )


def check_missing_created_date(rule, findings):
    if is_empty(get_value(rule, "Created_Date")):
        add_finding(
            rule, findings,
            "Missing created date",
            "Low",
            "Without a created date, it is difficult to determine the age of the rule "
            "or whether it predates current security policy.",
            "Add the date the rule was originally created."
        )


def check_stale_rule(rule, findings):
    """
    Flags rules not reviewed in over 365 days.
    """
    last_reviewed = get_value(rule, "Last_Reviewed_Date")

    if is_empty(last_reviewed):
        return

    parsed_date = pd.to_datetime(last_reviewed, dayfirst=True, errors="coerce")

    if pd.isna(parsed_date):
        return

    today = pd.Timestamp(date.today())
    days_since_review = (today - parsed_date.normalize()).days

    if days_since_review > 365:
        add_finding(
            rule, findings,
            f"Rule not reviewed in over 365 days ({days_since_review} days)",
            "Medium",
            f"This rule was last reviewed {days_since_review} days ago. Stale rules "
            "may no longer reflect current access requirements.",
            "Schedule a review of this rule to confirm whether it is still required."
        )


def check_allow_without_nat_in_external_rule(rule, findings):
    """
    Flags externally-facing allow rules that have no NAT configured.
    """
    action = normalize(get_value(rule, "Action"))
    source_zone = normalize(get_value(rule, "Source_Zone"))
    nat = normalize(get_value(rule, "NAT"))

    if (
        action == "ALLOW"
        and source_zone in EXTERNAL_ZONES
        and nat in {"NO", "NONE", "DISABLED", "FALSE", ""}
    ):
        add_finding(
            rule, findings,
            "External allow rule has no NAT configured",
            "Medium",
            "External-facing allow rules without NAT may expose internal IP addressing "
            "schemes to untrusted networks.",
            "Review whether NAT should be applied to mask internal addressing for "
            "this external-facing rule."
        )


# ============================================================
# Duplicate Rule Detection
# ============================================================

def check_duplicate_rules(df, findings):
    """
    Detects rules that share identical key traffic fields.
    Reports once per duplicate group with a count of matched rules.
    """
    duplicates = df[df.duplicated(subset=DUPLICATE_COLUMNS, keep=False)]

    if duplicates.empty:
        return

    groups = duplicates.groupby(DUPLICATE_COLUMNS, dropna=False)

    for _, group in groups:
        first_rule = group.iloc[0]
        rule_ids = group["Rule_ID"].tolist()
        count = len(rule_ids)

        add_finding(
            first_rule, findings,
            f"Possible duplicate rule detected ({count} matching rules: "
            f"{', '.join(str(r) for r in rule_ids)})",
            "Medium",
            "This rule shares the same source, destination, service, protocol, port, "
            "and action as one or more other rules. Duplicate rules make the rulebase "
            "harder to manage and audit.",
            "Review duplicate rules and remove unnecessary repeated entries."
        )


# ============================================================
# Shadowed Rule Detection
# ============================================================

def check_shadowed_rules(df, findings):
    """
    Detects rules that are shadowed by a preceding broader rule.

    A rule is considered shadowed when an earlier rule in the list
    has the same or broader source, destination, and service fields
    with the same action. The shadowed rule will never be evaluated
    by the firewall engine.

    This is a heuristic check based on ANY keyword matching.
    It does not perform full subnet or CIDR expansion.
    """
    for i, lower_rule in df.iterrows():
        lower_action = normalize(lower_rule.get("Action", ""))
        lower_source = normalize(lower_rule.get("Source", ""))
        lower_dest = normalize(lower_rule.get("Destination", ""))
        lower_service = normalize(lower_rule.get("Service", ""))

        for j, upper_rule in df.iloc[:i].iterrows():
            upper_action = normalize(upper_rule.get("Action", ""))
            upper_source = normalize(upper_rule.get("Source", ""))
            upper_dest = normalize(upper_rule.get("Destination", ""))
            upper_service = normalize(upper_rule.get("Service", ""))

            if upper_action != lower_action:
                continue

            source_shadowed = upper_source == "ANY" or upper_source == lower_source
            dest_shadowed = upper_dest == "ANY" or upper_dest == lower_dest
            service_shadowed = upper_service == "ANY" or upper_service == lower_service

            if source_shadowed and dest_shadowed and service_shadowed:
                add_finding(
                    lower_rule, findings,
                    f"Rule may be shadowed by Rule_ID {upper_rule.get('Rule_ID', j)}",
                    "Medium",
                    "A preceding rule with broader or identical match criteria may prevent "
                    "this rule from ever being evaluated by the firewall engine.",
                    "Review rule ordering and remove or reorder rules so that more specific "
                    "rules appear before broader ones."
                )
                break


# ============================================================
# Analyzer Engine
# ============================================================

def analyze_firewall_rules(df):
    """
    Runs all compliance checks against the loaded firewall rules.
    Returns a DataFrame of all findings.
    """
    findings = []

    validate_columns(df)

    for _, rule in df.iterrows():
        check_any_source(rule, findings)
        check_any_destination(rule, findings)
        check_any_service(rule, findings)
        check_any_any_any(rule, findings)
        check_missing_owner(rule, findings)
        check_missing_department(rule, findings)
        check_missing_business_justification(rule, findings)
        check_missing_ticket_number(rule, findings)
        check_missing_last_reviewed_date(rule, findings)
        check_missing_created_date(rule, findings)
        check_logging_disabled(rule, findings)
        check_rule_type_without_expiry(rule, findings)
        check_expired_rule(rule, findings)
        check_expiry_within_30_days(rule, findings)
        check_stale_rule(rule, findings)
        check_risky_services(rule, findings)
        check_disabled_rule(rule, findings)
        check_external_to_internal(rule, findings)
        check_prod_any_rule(rule, findings)
        check_intrazone_allow(rule, findings)
        check_allow_without_nat_in_external_rule(rule, findings)

    check_duplicate_rules(df, findings)
    check_shadowed_rules(df, findings)

    if not findings:
        return pd.DataFrame()

    findings_df = pd.DataFrame(findings)

    findings_df["Severity"] = pd.Categorical(
        findings_df["Severity"],
        categories=SEVERITY_ORDER,
        ordered=True
    )

    findings_df = findings_df.sort_values(
        ["Severity", "Rule_ID"],
        ascending=[True, True]
    ).reset_index(drop=True)

    return findings_df


# ============================================================
# Summary Builders
# ============================================================

def create_summary(findings_df):
    """
    Creates a severity count summary table.
    """
    if findings_df.empty:
        return pd.DataFrame({
            "Severity": SEVERITY_ORDER,
            "Count": [0, 0, 0, 0]
        })

    summary = findings_df["Severity"].value_counts().reset_index()
    summary.columns = ["Severity", "Count"]

    summary["Severity"] = pd.Categorical(
        summary["Severity"],
        categories=SEVERITY_ORDER,
        ordered=True
    )

    return summary.sort_values("Severity").reset_index(drop=True)


def create_rule_status_summary(original_df, findings_df):
    """
    Produces a per-rule pass/fail table with finding count and highest severity.
    Groups by row index to handle missing or duplicate Rule_IDs safely.
    """
    result_df = original_df[["Rule_ID", "Rule_Name"]].copy()
    result_df["Compliance_Status"] = "Pass"
    result_df["Finding_Count"] = 0
    result_df["Highest_Severity"] = "None"

    if findings_df.empty:
        return result_df

    for idx, row in result_df.iterrows():
        rule_id = row["Rule_ID"]
        rule_findings = findings_df[findings_df["Rule_ID"] == rule_id]

        if rule_findings.empty:
            continue

        result_df.at[idx, "Compliance_Status"] = "Fail"
        result_df.at[idx, "Finding_Count"] = len(rule_findings)
        result_df.at[idx, "Highest_Severity"] = rule_findings["Severity"].max()

    return result_df


# ============================================================
# Excel Styling Helpers
# ============================================================

def make_fill(hex_color):
    return PatternFill(
        fill_type="solid",
        fgColor=hex_color
    )


def make_font(bold=False, color="000000", size=11):
    return Font(
        bold=bold,
        color=color,
        size=size
    )


def make_alignment(wrap=True, horizontal="left", vertical="top"):
    return Alignment(
        wrap_text=wrap,
        horizontal=horizontal,
        vertical=vertical
    )


def make_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header_row(ws, header_fill_hex="1F3864"):
    """
    Styles the first row of a worksheet as a header.
    """
    header_fill = make_fill(header_fill_hex)
    header_font = make_font(bold=True, color="FFFFFF", size=11)
    center = make_alignment(wrap=False, horizontal="center", vertical="center")
    border = make_border()

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    ws.row_dimensions[1].height = 22


def style_data_rows(ws, severity_col_index=None):
    """
    Applies alternating row shading, borders, and optional severity
    color coding to data rows (row 2 onwards).

    severity_col_index: 1-based column index of the Severity column.
    """
    alt_fill = make_fill("F2F2F2")
    default_font = make_font()
    wrap_align = make_alignment()
    border = make_border()

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.font = default_font
            cell.alignment = wrap_align
            cell.border = border

            if row_idx % 2 == 0:
                cell.fill = alt_fill

        if severity_col_index:
            sev_cell = ws.cell(row=row_idx, column=severity_col_index)
            sev_value = str(sev_cell.value).strip() if sev_cell.value else ""

            if sev_value in SEVERITY_COLORS:
                color = SEVERITY_COLORS[sev_value]
                sev_cell.fill = make_fill(color)
                sev_cell.font = make_font(bold=True, color="FFFFFF")
                sev_cell.alignment = make_alignment(
                    horizontal="center",
                    vertical="top"
                )


def auto_fit_columns(ws, min_width=12, max_width=55):
    """
    Sets column widths based on the longest value in each column.
    """
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_length = 0

        for cell in col:
            try:
                cell_length = len(str(cell.value)) if cell.value else 0
                max_length = max(max_length, cell_length)
            except Exception:
                pass

        adjusted = min(max(max_length + 4, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def freeze_header(ws):
    ws.freeze_panes = "A2"


def add_autofilter(ws):
    ws.auto_filter.ref = ws.dimensions


def get_severity_col_index(ws):
    """
    Returns the 1-based column index of the Severity column, or None.
    """
    for cell in ws[1]:
        if str(cell.value).strip().lower() == "severity":
            return cell.column
    return None


def style_sheet(ws, severity_col_index=None):
    """
    Applies full styling to a worksheet.
    """
    style_header_row(ws)
    style_data_rows(ws, severity_col_index=severity_col_index)
    auto_fit_columns(ws)
    freeze_header(ws)
    add_autofilter(ws)


def style_compliance_status_column(ws):
    """
    Colors the Compliance_Status column green for Pass, red for Fail.
    """
    status_col = None

    for cell in ws[1]:
        if str(cell.value).strip().lower() == "compliance_status":
            status_col = cell.column
            break

    if not status_col:
        return

    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=status_col)
        value = str(cell.value).strip().upper() if cell.value else ""

        if value == "PASS":
            cell.fill = make_fill("70AD47")
            cell.font = make_font(bold=True, color="FFFFFF")
            cell.alignment = make_alignment(horizontal="center")
        elif value == "FAIL":
            cell.fill = make_fill("C00000")
            cell.font = make_font(bold=True, color="FFFFFF")
            cell.alignment = make_alignment(horizontal="center")


def write_summary_sheet(ws, summary_df):
    """
    Writes and styles the Summary sheet with severity color coding per row.
    """
    from openpyxl.utils.dataframe import dataframe_to_rows

    for row in dataframe_to_rows(summary_df, index=False, header=True):
        ws.append(row)

    style_header_row(ws)
    freeze_header(ws)
    auto_fit_columns(ws)

    for row_idx in range(2, ws.max_row + 1):
        sev_cell = ws.cell(row=row_idx, column=1)
        count_cell = ws.cell(row=row_idx, column=2)
        sev_value = str(sev_cell.value).strip() if sev_cell.value else ""

        border = make_border()

        for cell in [sev_cell, count_cell]:
            cell.border = border
            cell.alignment = make_alignment(horizontal="center")

        if sev_value in SEVERITY_COLORS:
            color = SEVERITY_COLORS[sev_value]
            sev_cell.fill = make_fill(color)
            sev_cell.font = make_font(bold=True, color="FFFFFF")
            count_cell.fill = make_fill(color)
            count_cell.font = make_font(bold=True, color="FFFFFF")


# ============================================================
# Report Export
# ============================================================

def export_report(original_df, findings_df, output_file):
    """
    Exports the full compliance report to a styled Excel workbook.

    Sheets:
      - Summary           : severity count table
      - Rule Status       : per-rule pass/fail with finding count
      - Findings          : all compliance findings
      - Original Rules    : the source data as loaded
    """
    summary_df = create_summary(findings_df)
    rule_status_df = create_rule_status_summary(original_df, findings_df)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

        # --- Summary sheet ---
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # --- Rule Status sheet ---
        rule_status_df.to_excel(writer, sheet_name="Rule Status", index=False)

        # --- Findings sheet ---
        if findings_df.empty:
            no_findings = pd.DataFrame({"Result": ["No compliance issues found."]})
            no_findings.to_excel(writer, sheet_name="Findings", index=False)
        else:
            findings_df.to_excel(writer, sheet_name="Findings", index=False)

        # --- Original Rules sheet ---
        original_df.to_excel(writer, sheet_name="Original Rules", index=False)

        workbook = writer.book

        # Style Summary
        ws_summary = workbook["Summary"]
        write_summary_sheet(ws_summary, summary_df)

        # Style Rule Status
        ws_status = workbook["Rule Status"]
        style_sheet(ws_status)
        style_compliance_status_column(ws_status)

        # Style Findings
        ws_findings = workbook["Findings"]
        sev_col = get_severity_col_index(ws_findings)
        style_sheet(ws_findings, severity_col_index=sev_col)

        # Style Original Rules
        ws_original = workbook["Original Rules"]
        style_sheet(ws_original)


# ============================================================
# CLI Entry Point
# ============================================================

def parse_args():
    parser = argparse.ArgumentParser(
        description="Firewall Compliance Analyzer",
        formatter_class=argparse.RawTextHelpFormatter
    )
    parser.add_argument(
        "input_file",
        nargs="?",
        default=None,
        help="Path to firewall rules file (CSV, XLSX, or XLS).\n"
             "If omitted, you will be prompted to enter it."
    )
    parser.add_argument(
        "-o", "--output",
        default="firewall_compliance_report.xlsx",
        help="Output Excel report filename.\n"
             "Default: firewall_compliance_report.xlsx"
    )
    parser.add_argument(
        "--no-shadow",
        action="store_true",
        help="Skip shadowed rule detection (faster on large rulebases)."
    )
    return parser.parse_args()


def main():
    args = parse_args()

    print("=" * 70)
    print("  Firewall Compliance Analyzer")
    print("=" * 70)

    file_path = args.input_file

    if not file_path:
        file_path = input("\nEnter firewall rule file path (CSV/XLSX): ").strip()

    try:
        print(f"\nLoading: {file_path}")
        df = load_firewall_rules(file_path)
        print(f"Rules loaded: {len(df)}")

        print("\nRunning compliance checks...")
        findings_df = analyze_firewall_rules(df)

        if args.no_shadow:
            print("Shadowed rule detection skipped (--no-shadow).")
        else:
            print("Running shadowed rule detection...")
            check_shadowed_rules(df, findings_df if not findings_df.empty else [])

        print("\nAnalysis complete.")

        if findings_df.empty:
            print("Result: No compliance issues found.")
        else:
            print(f"Total findings : {len(findings_df)}")
            print("\nFindings by severity:")
            print(create_summary(findings_df).to_string(index=False))

        print(f"\nExporting report to: {args.output}")
        export_report(df, findings_df, args.output)
        print(f"Report saved: {args.output}")

    except FileNotFoundError as e:
        print(f"\n[ERROR] {e}")
    except ValueError as e:
        print(f"\n[ERROR] {e}")
    except Exception as e:
        print(f"\n[UNEXPECTED ERROR] {e}")
        raise


if __name__ == "__main__":
    main()

